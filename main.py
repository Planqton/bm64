import asyncio
import os
from datetime import datetime

import config

from bleak import BleakClient, BleakScanner
from openpyxl import Workbook, load_workbook

# Konfiguration
DEVICE_ADDRESS = "A4:C1:38:A5:20:BB"  # <- anpassen, falls nötig
CHAR_UUID = "00002a35-0000-1000-8000-00805f9b34fb"
EXCEL_FILE = "log.xlsx"

# Robuster Parser nach Bluetooth-Spezifikation
def parse_measurement(data: bytes):
    try:
        flags = data[0]
        index = 1

        # Pflichtfelder (6 Bytes)
        systolic = int.from_bytes(data[index:index+2], byteorder='little')
        index += 2
        diastolic = int.from_bytes(data[index:index+2], byteorder='little')
        index += 2
        _ = int.from_bytes(data[index:index+2], byteorder='little')  # MAP ignorieren
        index += 2

        # Berechne MAP immer selbst
        map_val = int(diastolic + (systolic - diastolic) / 3)

        # Zeitstempel (Bit 0)
        if flags & 0x01:
            year = int.from_bytes(data[index:index+2], byteorder='little')
            month = data[index+2]
            day = data[index+3]
            hour = data[index+4]
            minute = data[index+5]
            second = data[index+6]
            timestamp = f"{year:04}-{month:02}-{day:02} {hour:02}:{minute:02}:{second:02}"
            index += 7
        else:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Puls (Bit 1)
        if flags & 0x02:
            pulse = int.from_bytes(data[index:index+2], byteorder='little')
            index += 2
        else:
            pulse = None

        # User ID (Bit 2) – optional
        if flags & 0x04:
            index += 1

        # Status (Bit 3) – optional
        if flags & 0x08:
            index += 2

        return {
            "timestamp": timestamp,
            "systole": systolic,
            "diastole": diastolic,
            "map": map_val,
            "pulse": pulse
        }

    except Exception as e:
        print(f"❌ Parser-Fehler: {e}")
        return None

# Excel-Schreiber
def write_to_excel(values):
    path = os.path.join(os.path.dirname(__file__), EXCEL_FILE)
    headers = ["Zeitstempel", "Systole", "Diastole", "MAP", "Puls"]

    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Messwerte"
        ws.append(headers)
    else:
        wb = load_workbook(path)
        ws = wb.active

    ws.append([
        values["timestamp"],
        values["systole"],
        values["diastole"],
        values["map"],
        values["pulse"] if values["pulse"] is not None else ""
    ])

    wb.save(path)
    print(f"✅ Gespeichert: {values}")

# BLE-Callback bei Benachrichtigung
def handle_notification(sender, data):
    print(f"\n🔵 Empfangene Daten von {sender}")
    print(f"Raw: {data.hex()}")
    values = parse_measurement(data)
    if values:
        write_to_excel(values)


async def fetch_loop(address):
    async with BleakClient(address) as client:
        if not client.is_connected:
            await client.connect()
        print("✅ Verbunden mit Blutdruckmessgerät")
        await client.start_notify(CHAR_UUID, handle_notification)
        try:
            while True:
                await asyncio.sleep(1)
        except asyncio.CancelledError:
            await client.stop_notify(CHAR_UUID)
            print("⏹️ Fetching gestoppt")
            raise


async def pair_device(address: str) -> bool:
    try:
        async with BleakClient(address) as client:
            if not client.is_connected:
                await client.connect()
            paired = await client.pair()
            if paired:
                print(f"✅ Gepaart mit {address}")
            else:
                print(f"❌ Pairing mit {address} fehlgeschlagen")
            return paired
    except Exception as e:
        print(f"❌ Pairing-Fehler: {e}")
        return False


async def configure(cfg):
    print("\n🔍 Suche Bluetooth-Geräte...")
    devices = await BleakScanner.discover(timeout=5.0)
    if not devices:
        print("Keine Geräte gefunden")
        return
    for idx, d in enumerate(devices, 1):
        name = d.name or "Unbekannt"
        print(f"{idx}) {name} [{d.address}]")
    choice = input("Gerät wählen (Nummer) oder Enter abbrechen: ").strip()
    if choice.isdigit():
        i = int(choice) - 1
        if 0 <= i < len(devices):
            address = devices[i].address
            print(f"🔗 Versuche Pairing mit {address} ...")
            if await pair_device(address):
                cfg["device_address"] = address
                config.save_config(cfg)
                print(f"Gerät {address} gespeichert")
        else:
            print("Ungültige Auswahl")

# Haupt-Async-Logik für interaktive Bedienung
async def main():
    cfg = config.load_config()
    fetch_task = None

    while True:
        running = fetch_task is not None and not fetch_task.done()
        status = "running" if running else "not run"
        print(f"\nFetching ({status})")
        print("1) Configuration")
        print("2) Stop Fetching" if running else "2) Start Fetching")
        print("3) Exit")
        choice = input("> ").strip()

        if choice == "1":
            await configure(cfg)
        elif choice == "2":
            if running:
                fetch_task.cancel()
                try:
                    await fetch_task
                except asyncio.CancelledError:
                    pass
                fetch_task = None
            else:
                address = cfg.get("device_address")
                if not address:
                    print("Kein Gerät konfiguriert.")
                else:
                    fetch_task = asyncio.create_task(fetch_loop(address))
        elif choice == "3":
            if running:
                fetch_task.cancel()
                try:
                    await fetch_task
                except asyncio.CancelledError:
                    pass
            break
        else:
            print("Ungültige Eingabe")

# Einstiegspunkt
if __name__ == "__main__":
    asyncio.run(main())
